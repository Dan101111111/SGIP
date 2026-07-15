from fastapi import APIRouter, Depends, Query, HTTPException
from datetime import datetime, timedelta
from typing import Optional
from app.services.report_service import ReportService
from app.schemas.report_schema import DailyReportResponse, WeeklyReportResponse
from app.core.exceptions import NotFoundException
from app.core.config import settings

router = APIRouter(prefix="/api/reports", tags=["Reports"])


@router.get("/daily")
async def get_daily_report(
    date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format"),
    service: ReportService = Depends()
) -> DailyReportResponse:
    """Get daily report"""
    if date:
        try:
            report_date = datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        report_date = None
    
    report = service.generate_daily_report(report_date)
    return DailyReportResponse(**report)


@router.get("/moche/daily")
async def get_moche_daily_report(
    date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format"),
    service: ReportService = Depends()
) -> DailyReportResponse:
    """Get daily report for Moche sector"""
    if date:
        try:
            report_date = datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        report_date = None
    
    report = service.generate_daily_report(report_date)
    # Ensure it's for Moche
    if report["dma"]["code"] != settings.target_dma:
        raise NotFoundException("Report", f"Moche sector on {date or 'today'}")
    return DailyReportResponse(**report)


@router.get("/weekly")
async def get_weekly_report(
    service: ReportService = Depends()
) -> WeeklyReportResponse:
    """Get weekly report"""
    report = service.generate_weekly_report()
    return WeeklyReportResponse(**report)


@router.get("/moche/weekly")
async def get_moche_weekly_report(
    service: ReportService = Depends()
) -> WeeklyReportResponse:
    """Get weekly report for Moche sector"""
    report = service.generate_weekly_report()
    if report["dma"] != settings.target_dma:
        raise NotFoundException("Report", "Moche sector")
    return WeeklyReportResponse(**report)


@router.get("/custom")
async def get_custom_report(
    start_date: datetime = Query(..., description="Start date"),
    end_date: datetime = Query(..., description="End date"),
    dma_id: Optional[str] = Query(None, description="DMA ID"),
    service: ReportService = Depends()
):
    """Get custom report for a date range"""
    if start_date > end_date:
        raise HTTPException(status_code=400, detail="Start date must be before end date")
    
    if (end_date - start_date).days > 90:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 90 days")
    
    # Get data for the period
    dma = dma_id or settings.target_dma
    
    readings = service.telemetry_service.get_historical_readings(
        dma,
        start_date,
        end_date,
        limit=10000
    )
    
    anomalies = service.anomaly_service.get_recent_anomalies(
        dma,
        hours=int((end_date - start_date).total_seconds() / 3600)
    )
    
    incidents = service.incident_service.get_all_tickets(
        dma_id=dma,
        limit=1000
    )
    
    return {
        "period": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat()
        },
        "dma_id": dma,
        "statistics": {
            "total_readings": len(readings),
            "avg_pressure": round(sum(r.pressure_mca for r in readings) / len(readings), 1) if readings else 0,
            "avg_flow": round(sum(r.flow_lps for r in readings) / len(readings), 1) if readings else 0,
            "anomalies_detected": len(anomalies),
            "incidents_created": len(incidents),
            "incidents_resolved": len([i for i in incidents if i.status.value in ["RESOLVED", "CLOSED"]])
        },
        "generated_at": datetime.utcnow().isoformat()
    }


@router.get("/moche/custom")
async def get_moche_custom_report(
    start_date: datetime = Query(..., description="Start date"),
    end_date: datetime = Query(..., description="End date"),
    service: ReportService = Depends()
):
    """Get custom report for Moche sector"""
    return await get_custom_report(start_date, end_date, settings.target_dma, service)


@router.get("/export/{report_type}")
async def export_report(
    report_type: str,
    format: str = Query("json", description="Export format (json, csv, or pdf)"),
    date: Optional[str] = Query(None, description="Date for daily report"),
    start_date: Optional[datetime] = Query(None, description="Start date for custom report"),
    end_date: Optional[datetime] = Query(None, description="End date for custom report"),
    service: ReportService = Depends()
):
    """Export a report in specified format"""
    if report_type not in ["daily", "weekly", "custom"]:
        raise HTTPException(status_code=400, detail="Invalid report type. Use 'daily', 'weekly' or 'custom'")
    
    if report_type == "daily":
        report_date = None
        if date:
            try:
                report_date = datetime.strptime(date, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        report = service.generate_daily_report(report_date)
        title = f"Reporte Diario - {report.get('date', '')}"
    elif report_type == "weekly":
        report = service.generate_weekly_report()
        title = f"Reporte Semanal"
    else:
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="start_date and end_date are required for custom report")
        
        dma = settings.target_dma
        readings = service.telemetry_service.get_historical_readings(dma, start_date, end_date, limit=10000)
        anomalies = service.anomaly_service.get_recent_anomalies(dma, hours=int((end_date - start_date).total_seconds() / 3600))
        incidents = service.incident_service.get_all_tickets(dma_id=dma, limit=1000)
        
        report = {
            "period": {"start": start_date.isoformat(), "end": end_date.isoformat()},
            "dma_id": dma,
            "statistics": {
                "total_readings": len(readings),
                "avg_pressure": round(sum(r.pressure_mca for r in readings) / len(readings), 1) if readings else 0,
                "avg_flow": round(sum(r.flow_lps for r in readings) / len(readings), 1) if readings else 0,
                "anomalies_detected": len(anomalies),
                "incidents_created": len(incidents),
                "incidents_resolved": len([i for i in incidents if i.status.value in ["RESOLVED", "CLOSED"]])
            },
            "anomalies_list": [{"id": a.get("anomaly").id if a.get("anomaly") else a.get("id", "N/A"), "date": (a.get("anomaly").detected_at.isoformat() if a.get("anomaly") and a.get("anomaly").detected_at else a.get("detected_at", "")), "severity": (a.get("anomaly").severity.value if a.get("anomaly") and hasattr(a.get("anomaly").severity, "value") else a.get("severity", "")), "status": (a.get("anomaly").status.value if a.get("anomaly") and hasattr(a.get("anomaly").status, "value") else a.get("status", ""))} for a in anomalies],
            "incidents_list": [{"code": i.code, "title": i.title, "priority": i.priority.value if hasattr(i.priority, 'value') else i.priority, "status": i.status.value if hasattr(i.status, 'value') else i.status, "date": i.created_at.isoformat()} for i in incidents],
            "generated_at": datetime.utcnow().isoformat()
        }
        title = f"Reporte Personalizado: {start_date.strftime('%Y-%m-%d')} a {end_date.strftime('%Y-%m-%d')}"
    
    if format.lower() == "csv":
        import csv
        from io import StringIO
        output = StringIO()
        writer = csv.writer(output)
        
        writer.writerow(["SISTEMA DE GESTIÓN INTEGRAL DE PÉRDIDAS (SGIP)"])
        writer.writerow([title])
        writer.writerow(["Generado el:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")])
        writer.writerow([])
        
        if report_type == "daily":
            writer.writerow(["RESUMEN EJECUTIVO"])
            writer.writerow(["Métrica", "Valor"])
            writer.writerow(["DMA", report.get("dma", {}).get("name")])
            writer.writerow(["Presión Promedio (MCA)", report.get("summary", {}).get("avg_pressure")])
            writer.writerow(["Caudal Promedio (LPS)", report.get("summary", {}).get("avg_flow")])
            writer.writerow(["Anomalías Totales", report.get("anomalies", {}).get("total")])
            writer.writerow(["Incidentes Totales", report.get("incidents", {}).get("total")])
            
            writer.writerow([])
            writer.writerow(["DETALLES DE LECTURAS"])
            writer.writerow(["Hora", "Presión (MCA)", "Caudal (LPS)", "Anomalía"])
            for r in report.get("readings", []):
                time_str = datetime.fromisoformat(r["timestamp"]).strftime("%H:%M:%S")
                writer.writerow([time_str, r["pressure_mca"], r["flow_lps"], "Sí" if r.get("is_anomaly") else "No"])
                
        elif report_type == "weekly":
            writer.writerow(["RESUMEN EJECUTIVO"])
            writer.writerow(["Métrica", "Valor"])
            writer.writerow(["DMA", report.get("dma")])
            writer.writerow(["Lecturas Totales", report.get("total_readings")])
            writer.writerow(["Anomalías Totales", report.get("total_anomalies")])
            writer.writerow(["Incidentes Totales", report.get("total_incidents")])
            writer.writerow(["Pérdida de Agua (m3)", report.get("water_loss_estimate")])
            
            writer.writerow([])
            writer.writerow(["ESTADÍSTICAS DIARIAS"])
            writer.writerow(["Fecha", "Presión Promedio (MCA)", "Caudal Promedio (LPS)", "Anomalías", "Pérdida Est. (m3)"])
            for day in report.get("daily_stats", []):
                writer.writerow([day.get("date"), day.get("avg_pressure"), day.get("avg_flow"), day.get("anomalies_count"), day.get("water_loss")])
                
        else: # custom
            writer.writerow(["RESUMEN EJECUTIVO"])
            writer.writerow(["Métrica", "Valor"])
            writer.writerow(["DMA", report.get("dma_id")])
            stats = report.get("statistics", {})
            writer.writerow(["Lecturas Totales", stats.get("total_readings")])
            writer.writerow(["Presión Promedio (MCA)", stats.get("avg_pressure")])
            writer.writerow(["Caudal Promedio (LPS)", stats.get("avg_flow")])
            writer.writerow(["Anomalías Detectadas", stats.get("anomalies_detected")])
            writer.writerow(["Incidentes Creados", stats.get("incidents_created")])
            writer.writerow(["Incidentes Resueltos", stats.get("incidents_resolved")])
            
            writer.writerow([])
            writer.writerow(["DETALLES DE ANOMALÍAS"])
            writer.writerow(["ID", "Fecha", "Severidad", "Estado"])
            for a in report.get("anomalies_list", []):
                writer.writerow([a.get("id"), a.get("date"), a.get("severity"), a.get("status")])
                
            writer.writerow([])
            writer.writerow(["DETALLES DE INCIDENTES"])
            writer.writerow(["Código", "Título", "Prioridad", "Estado", "Fecha"])
            for i in report.get("incidents_list", []):
                writer.writerow([i.get("code"), i.get("title"), i.get("priority"), i.get("status"), i.get("date")])
            
        import base64
        return {
            "content": base64.b64encode(output.getvalue().encode('utf-8')).decode('utf-8'),
            "filename": f"{report_type}_report_{datetime.utcnow().strftime('%Y%m%d')}.csv",
            "format": "csv"
        }
        
    elif format.lower() == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        import base64
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="0D6EBD")
        title_font = Font(bold=True, size=14)
        
        ws.append(["SISTEMA DE GESTIÓN INTEGRAL DE PÉRDIDAS (SGIP)"])
        ws["A1"].font = title_font
        ws.append([title])
        ws.append(["Generado el:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])
        
        if report_type == "daily":
            ws.append(["RESUMEN EJECUTIVO"])
            ws.cell(row=5, column=1).font = Font(bold=True)
            
            headers = ["Métrica", "Valor"]
            ws.append(headers)
            for cell in ws[6]:
                cell.font = header_font
                cell.fill = header_fill
            
            ws.append(["DMA", report.get("dma", {}).get("name")])
            ws.append(["Presión Promedio (MCA)", report.get("summary", {}).get("avg_pressure")])
            ws.append(["Caudal Promedio (LPS)", report.get("summary", {}).get("avg_flow")])
            ws.append(["Anomalías Totales", report.get("anomalies", {}).get("total")])
            ws.append(["Incidentes Totales", report.get("incidents", {}).get("total")])
            
            ws.append([])
            ws.append(["DETALLES DE LECTURAS"])
            ws.cell(row=13, column=1).font = Font(bold=True)
            
            headers = ["Hora", "Presión (MCA)", "Caudal (LPS)", "Anomalía"]
            ws.append(headers)
            for cell in ws[14]:
                cell.font = header_font
                cell.fill = header_fill
                
            for r in report.get("readings", []):
                time_str = datetime.fromisoformat(r["timestamp"]).strftime("%H:%M:%S")
                ws.append([time_str, r["pressure_mca"], r["flow_lps"], "Sí" if r.get("is_anomaly") else "No"])
                if r.get("is_anomaly"):
                    ws.cell(row=ws.max_row, column=4).font = Font(color="FF0000")
                    
        elif report_type == "weekly":
            ws.append(["RESUMEN EJECUTIVO"])
            ws.cell(row=5, column=1).font = Font(bold=True)
            
            headers = ["Métrica", "Valor"]
            ws.append(headers)
            for cell in ws[6]:
                cell.font = header_font
                cell.fill = header_fill
                
            ws.append(["DMA", report.get("dma")])
            ws.append(["Lecturas Totales", report.get("total_readings")])
            ws.append(["Anomalías Totales", report.get("total_anomalies")])
            ws.append(["Incidentes Totales", report.get("total_incidents")])
            ws.append(["Pérdida de Agua (m3)", report.get("water_loss_estimate")])
            
            ws.append([])
            ws.append(["ESTADÍSTICAS DIARIAS"])
            ws.cell(row=13, column=1).font = Font(bold=True)
            
            headers = ["Fecha", "Presión Promedio (MCA)", "Caudal Promedio (LPS)", "Anomalías", "Pérdida Est. (m3)"]
            ws.append(headers)
            for cell in ws[14]:
                cell.font = header_font
                cell.fill = header_fill
                
            for day in report.get("daily_stats", []):
                ws.append([day.get("date"), day.get("avg_pressure"), day.get("avg_flow"), day.get("anomalies_count"), day.get("water_loss")])
                
        else: # custom
            ws.append(["RESUMEN EJECUTIVO"])
            ws.cell(row=5, column=1).font = Font(bold=True)
            
            headers = ["Métrica", "Valor"]
            ws.append(headers)
            for cell in ws[6]:
                cell.font = header_font
                cell.fill = header_fill
                
            ws.append(["DMA", report.get("dma_id")])
            stats = report.get("statistics", {})
            ws.append(["Lecturas Totales", stats.get("total_readings")])
            ws.append(["Presión Promedio (MCA)", stats.get("avg_pressure")])
            ws.append(["Caudal Promedio (LPS)", stats.get("avg_flow")])
            ws.append(["Anomalías Detectadas", stats.get("anomalies_detected")])
            ws.append(["Incidentes Creados", stats.get("incidents_created")])
            ws.append(["Incidentes Resueltos", stats.get("incidents_resolved")])
            
            ws.append([])
            ws.append(["DETALLES DE ANOMALÍAS"])
            r = ws.max_row
            ws.cell(row=r, column=1).font = Font(bold=True)
            headers = ["ID", "Fecha", "Severidad", "Estado"]
            ws.append(headers)
            for cell in ws[r+1]:
                cell.font = header_font
                cell.fill = header_fill
            for a in report.get("anomalies_list", []):
                ws.append([a.get("id"), a.get("date"), a.get("severity"), a.get("status")])
                
            ws.append([])
            ws.append(["DETALLES DE INCIDENTES"])
            r = ws.max_row
            ws.cell(row=r, column=1).font = Font(bold=True)
            headers = ["Código", "Título", "Prioridad", "Estado", "Fecha"]
            ws.append(headers)
            for cell in ws[r+1]:
                cell.font = header_font
                cell.fill = header_fill
            for i in report.get("incidents_list", []):
                ws.append([i.get("code"), i.get("title"), i.get("priority"), i.get("status"), i.get("date")])
            
        output = BytesIO()
        wb.save(output)
        
        return {
            "content": base64.b64encode(output.getvalue()).decode('utf-8'),
            "filename": f"{report_type}_report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
            "format": "xlsx"
        }
        
    elif format.lower() == "pdf":
        from fpdf import FPDF
        import base64
        
        class PDFReport(FPDF):
            def header(self):
                self.set_font("helvetica", "B", 16)
                self.set_text_color(13, 110, 189) # Primary color
                self.cell(0, 10, "SGIP - Sistema de Gestión Integral de Pérdidas", ln=True, align="C")
                self.set_font("helvetica", "B", 12)
                self.set_text_color(50, 50, 50)
                self.cell(0, 10, title, ln=True, align="C")
                self.ln(5)
                
            def footer(self):
                self.set_y(-15)
                self.set_font("helvetica", "I", 8)
                self.set_text_color(128)
                self.cell(0, 10, f"Página {self.page_no()}", align="C")
                
            def chapter_title(self, title):
                self.set_font("helvetica", "B", 12)
                self.set_fill_color(240, 245, 255)
                self.set_text_color(13, 110, 189)
                self.cell(0, 8, title, ln=True, fill=True)
                self.ln(4)
                
            def metric_row(self, label, value):
                self.set_font("helvetica", "B", 10)
                self.set_text_color(80, 80, 80)
                self.cell(80, 7, label, border="B")
                self.set_font("helvetica", "", 10)
                self.set_text_color(0, 0, 0)
                self.cell(0, 7, str(value), border="B", ln=True)

        pdf = PDFReport()
        pdf.add_page()
        
        pdf.set_font("helvetica", "", 10)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, f"Generado el: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
        pdf.ln(5)
        
        pdf.chapter_title("Resumen Ejecutivo")
        
        if report_type == "daily":
            pdf.metric_row("Sector DMA", report.get("dma", {}).get("name", "N/A"))
            pdf.metric_row("Presión Promedio", f"{report.get('summary', {}).get('avg_pressure', 0)} MCA")
            pdf.metric_row("Caudal Promedio", f"{report.get('summary', {}).get('avg_flow', 0)} LPS")
            pdf.metric_row("Total de Anomalías", report.get("anomalies", {}).get("total", 0))
            pdf.metric_row("Total de Incidentes", report.get("incidents", {}).get("total", 0))
            
            pdf.ln(10)
            pdf.chapter_title("Detalles de Lecturas Recientes")
            
            pdf.set_font("helvetica", "B", 9)
            pdf.set_fill_color(220, 220, 220)
            pdf.cell(40, 7, "Hora", border=1, fill=True, align="C")
            pdf.cell(50, 7, "Presión (MCA)", border=1, fill=True, align="C")
            pdf.cell(50, 7, "Caudal (LPS)", border=1, fill=True, align="C")
            pdf.cell(40, 7, "Estado", border=1, fill=True, align="C", ln=True)
            
            pdf.set_font("helvetica", "", 9)
            for r in report.get("readings", [])[-20:]: # Show last 20
                time_str = datetime.fromisoformat(r["timestamp"]).strftime("%H:%M:%S")
                pdf.cell(40, 7, time_str, border=1, align="C")
                pdf.cell(50, 7, str(r.get("pressure_mca", 0)), border=1, align="C")
                pdf.cell(50, 7, str(r.get("flow_lps", 0)), border=1, align="C")
                status = "Anomalía" if r.get("is_anomaly") else "Normal"
                if r.get("is_anomaly"):
                    pdf.set_text_color(200, 0, 0)
                else:
                    pdf.set_text_color(0, 150, 0)
                pdf.cell(40, 7, status, border=1, align="C", ln=True)
                pdf.set_text_color(0, 0, 0)
                
        elif report_type == "weekly":
            pdf.metric_row("Sector DMA", report.get("dma", "N/A"))
            pdf.metric_row("Total Lecturas", report.get("total_readings", 0))
            pdf.metric_row("Anomalías Detectadas", report.get("total_anomalies", 0))
            pdf.metric_row("Incidentes Registrados", report.get("total_incidents", 0))
            pdf.metric_row("Pérdida Estimada de Agua", f"{report.get('water_loss_estimate', 0)} m³")
            
            pdf.ln(10)
            pdf.chapter_title("Estadísticas Diarias")
            
            pdf.set_font("helvetica", "B", 9)
            pdf.set_fill_color(220, 220, 220)
            pdf.cell(35, 7, "Fecha", border=1, fill=True, align="C")
            pdf.cell(35, 7, "Presión (MCA)", border=1, fill=True, align="C")
            pdf.cell(35, 7, "Caudal (LPS)", border=1, fill=True, align="C")
            pdf.cell(35, 7, "Anomalías", border=1, fill=True, align="C")
            pdf.cell(40, 7, "Pérdida (m³)", border=1, fill=True, align="C", ln=True)
            
            pdf.set_font("helvetica", "", 9)
            for day in report.get("daily_stats", []):
                pdf.cell(35, 7, str(day.get("date", "")), border=1, align="C")
                pdf.cell(35, 7, str(day.get("avg_pressure", 0)), border=1, align="C")
                pdf.cell(35, 7, str(day.get("avg_flow", 0)), border=1, align="C")
                pdf.cell(35, 7, str(day.get("anomalies_count", 0)), border=1, align="C")
                pdf.cell(40, 7, str(day.get("water_loss", 0)), border=1, align="C", ln=True)
                
        else: # custom
            stats = report.get("statistics", {})
            pdf.metric_row("Sector DMA", report.get("dma_id", "N/A"))
            pdf.metric_row("Total Lecturas", stats.get("total_readings", 0))
            pdf.metric_row("Presión Promedio", f"{stats.get('avg_pressure', 0)} MCA")
            pdf.metric_row("Caudal Promedio", f"{stats.get('avg_flow', 0)} LPS")
            pdf.metric_row("Anomalías Detectadas", stats.get("anomalies_detected", 0))
            pdf.metric_row("Incidentes Creados", stats.get("incidents_created", 0))
            pdf.metric_row("Incidentes Resueltos", stats.get("incidents_resolved", 0))
            
            pdf.ln(10)
            pdf.chapter_title("Detalles de Anomalías")
            pdf.set_font("helvetica", "B", 9)
            pdf.set_fill_color(220, 220, 220)
            pdf.cell(30, 7, "ID", border=1, fill=True, align="C")
            pdf.cell(50, 7, "Fecha", border=1, fill=True, align="C")
            pdf.cell(40, 7, "Severidad", border=1, fill=True, align="C")
            pdf.cell(40, 7, "Estado", border=1, fill=True, align="C", ln=True)
            pdf.set_font("helvetica", "", 9)
            for a in report.get("anomalies_list", []):
                pdf.cell(30, 7, str(a.get("id", "")), border=1, align="C")
                pdf.cell(50, 7, str(a.get("date", "")).split("T")[0] if "T" in str(a.get("date", "")) else str(a.get("date", "")), border=1, align="C")
                pdf.cell(40, 7, str(a.get("severity", "")), border=1, align="C")
                pdf.cell(40, 7, str(a.get("status", "")), border=1, align="C", ln=True)
                
            pdf.ln(10)
            pdf.chapter_title("Detalles de Incidentes")
            pdf.set_font("helvetica", "B", 9)
            pdf.set_fill_color(220, 220, 220)
            pdf.cell(35, 7, "Código", border=1, fill=True, align="C")
            pdf.cell(65, 7, "Título", border=1, fill=True, align="C")
            pdf.cell(30, 7, "Prioridad", border=1, fill=True, align="C")
            pdf.cell(30, 7, "Estado", border=1, fill=True, align="C", ln=True)
            pdf.set_font("helvetica", "", 9)
            for i in report.get("incidents_list", []):
                pdf.cell(35, 7, str(i.get("code", "")), border=1, align="C")
                pdf.cell(65, 7, str(i.get("title", ""))[:30], border=1, align="C")
                pdf.cell(30, 7, str(i.get("priority", "")), border=1, align="C")
                pdf.cell(30, 7, str(i.get("status", "")), border=1, align="C", ln=True)
            
        pdf_bytes = pdf.output(dest='S')
        
        return {
            "content": base64.b64encode(pdf_bytes).decode('utf-8'),
            "filename": f"{report_type}_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf",
            "format": "pdf"
        }
    
    return {
        "report": report,
        "format": "json",
        "generated_at": datetime.utcnow().isoformat()
    }